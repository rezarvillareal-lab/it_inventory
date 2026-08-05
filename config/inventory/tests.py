import datetime
import io
import csv

from django.contrib.auth import get_user_model
from django.test import TestCase
from django.urls import reverse

from .models import EquipmentComponent, Inventory, UserProfile


class RootViewTests(TestCase):
    def test_anonymous_user_is_redirected_to_login_from_home(self):
        response = self.client.get("/")

        self.assertEqual(response.status_code, 302)
        self.assertRedirects(response, reverse("login"))


class DashboardViewTests(TestCase):
    def _create_inventory(self, **overrides):
        defaults = {
            "control_number": "CN-0001",
            "office_or_hospital": "Office A",
            "user_name": "User 1",
            "computer_name": "PC-1",
            "assigned_ip": "10.0.0.1",
            "received_by": "Receiver",
            "position": "Staff",
            "date_received": datetime.date(2026, 3, 1),
            "created_at": datetime.date(2026, 3, 19),
            "status": Inventory.Status.ACTIVE,
        }
        defaults.update(overrides)
        return Inventory.objects.create(**defaults)

    def test_dashboard_renders_both_charts_and_context(self):
        user = get_user_model().objects.create_superuser(username="admin", password="password", email="admin@example.com")
        self.client.force_login(user)

        self._create_inventory(control_number="CN-1", office_or_hospital="Office A", status=Inventory.Status.ACTIVE)
        self._create_inventory(control_number="CN-2", office_or_hospital="Office A", status=Inventory.Status.ACTIVE)
        self._create_inventory(control_number="CN-3", office_or_hospital="Office B", status=Inventory.Status.MAINTENANCE)
        self._create_inventory(control_number="CN-4", office_or_hospital="Office B", status=Inventory.Status.CONDEMNED)
        self._create_inventory(control_number="CN-5", office_or_hospital="Office C", status=Inventory.Status.DISPOSED)

        response = self.client.get(reverse("inventory:dashboard"))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'id="officeChart"')
        self.assertContains(response, 'id="statusChart"')
        self.assertContains(response, 'id="equipmentChart"')
        self.assertContains(response, 'Equipment Summary')

        office_labels = response.context["office_labels"]
        office_counts = response.context["office_counts"]
        status_labels = response.context["status_labels"]
        status_counts = response.context["status_counts"]
        self.assertIn("total_equipment_components", response.context)
        self.assertIn("equipment_type_count", response.context)
        self.assertIn("top_equipment_breakdown", response.context)
        self.assertIn("equipment_labels", response.context)
        self.assertIn("equipment_counts", response.context)

        self.assertIsInstance(office_labels, list)
        self.assertIsInstance(office_counts, list)
        self.assertIsInstance(status_labels, list)
        self.assertIsInstance(status_counts, list)

        office_totals = dict(zip(office_labels, office_counts))
        self.assertEqual(office_totals.get("Office A"), 2)
        self.assertEqual(office_totals.get("Office B"), 2)
        self.assertEqual(office_totals.get("Office C"), 1)

        status_totals = dict(zip(status_labels, status_counts))
        self.assertEqual(status_totals.get(Inventory.Status.ACTIVE), 2)
        self.assertEqual(status_totals.get(Inventory.Status.MAINTENANCE), 1)
        self.assertEqual(status_totals.get(Inventory.Status.CONDEMNED), 1)
        self.assertEqual(status_totals.get(Inventory.Status.DISPOSED), 1)

    def test_dashboard_equipment_summary_includes_component_counts(self):
        user = get_user_model().objects.create_superuser(username="admin", password="password", email="admin@example.com")
        self.client.force_login(user)

        inventory1 = self._create_inventory(control_number="CN-1", office_or_hospital="Office A")
        inventory2 = self._create_inventory(control_number="CN-2", office_or_hospital="Office A")

        EquipmentComponent.objects.create(
            inventory=inventory1,
            component_name="Processor",
            original_model="Intel i7",
            original_serial="PROC-123",
            replacement_model="",
            replacement_serial="",
            remarks="",
        )
        EquipmentComponent.objects.create(
            inventory=inventory1,
            component_name="Monitor",
            original_model="Dell",
            original_serial="MON-456",
            replacement_model="",
            replacement_serial="",
            remarks="",
        )
        EquipmentComponent.objects.create(
            inventory=inventory2,
            component_name="Processor",
            original_model="Intel i5",
            original_serial="PROC-789",
            replacement_model="",
            replacement_serial="",
            remarks="",
        )

        response = self.client.get(reverse("inventory:dashboard"))

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.context["total_equipment_components"], 3)
        self.assertEqual(response.context["equipment_type_count"], 2)
        self.assertEqual(response.context["top_equipment_breakdown"][0]["component_name"], "Processor")
        self.assertEqual(response.context["top_equipment_breakdown"][0]["count"], 2)
        self.assertEqual(response.context["top_equipment_breakdown"][1]["component_name"], "Monitor")
        self.assertEqual(response.context["top_equipment_breakdown"][1]["count"], 1)

    def test_dashboard_shows_only_items_for_the_user_office_scope(self):
        user = get_user_model().objects.create_user(username="office_a_user", password="password")
        UserProfile.objects.create(user=user, office_or_hospital="Office A")
        self.client.force_login(user)

        self._create_inventory(control_number="CN-1", office_or_hospital="Office A", status=Inventory.Status.ACTIVE)
        self._create_inventory(control_number="CN-2", office_or_hospital="Office B", status=Inventory.Status.MAINTENANCE)

        response = self.client.get(reverse("inventory:dashboard"))

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.context["total_computers"], 1)
        self.assertEqual(response.context["office_labels"], ["Office A"])
        self.assertEqual(response.context["office_counts"], [1])
        self.assertEqual(response.context["status_labels"], [Inventory.Status.ACTIVE])
        self.assertEqual(response.context["status_counts"], [1])


class InventorySearchTests(TestCase):
    def _create_inventory(self, **overrides):
        defaults = {
            "control_number": "CN-SEARCH",
            "office_or_hospital": "Office A",
            "user_name": "User 1",
            "computer_name": "PC-1",
            "assigned_ip": "10.0.0.1",
            "received_by": "Receiver",
            "position": "Staff",
            "date_received": datetime.date(2026, 3, 1),
            "created_at": datetime.date(2026, 3, 19),
            "status": Inventory.Status.ACTIVE,
        }
        defaults.update(overrides)
        return Inventory.objects.create(**defaults)

    def test_inventory_list_exposes_equipment_filter_options(self):
        inventory = self._create_inventory(control_number="CN-EQ-1")
        EquipmentComponent.objects.create(
            inventory=inventory,
            component_name="Processor",
            original_model="Intel i7",
            original_serial="PROC-123",
            replacement_model="",
            replacement_serial="",
            remarks="",
        )

        user = get_user_model().objects.create_user(username="filter_user", password="password")
        UserProfile.objects.create(user=user, office_or_hospital="Office A")
        self.client.force_login(user)

        response = self.client.get(reverse("inventory:inventory_list"))

        self.assertEqual(response.status_code, 200)
        self.assertIn("equipment_options", response.context)
        self.assertTrue(
            any(option.get("value") == "Processor" for option in response.context["equipment_options"])
        )

    def test_inventory_search_matches_equipment_component_details(self):
        inventory = self._create_inventory(control_number="CN-EQ-1")
        other_inventory = self._create_inventory(control_number="CN-EQ-2")

        EquipmentComponent.objects.create(
            inventory=inventory,
            component_name="Processor",
            original_model="Intel i7",
            original_serial="PROC-123",
            replacement_model="",
            replacement_serial="",
            remarks="",
        )
        EquipmentComponent.objects.create(
            inventory=other_inventory,
            component_name="Monitor",
            original_model="Dell 24",
            original_serial="MON-001",
            replacement_model="",
            replacement_serial="",
            remarks="",
        )

        user = get_user_model().objects.create_user(username="search_user", password="password")
        UserProfile.objects.create(user=user, office_or_hospital="Office A")
        self.client.force_login(user)

        response = self.client.get(reverse("inventory:inventory_list") + "?q=Intel")

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "CN-EQ-1")
        self.assertNotContains(response, "CN-EQ-2")


class ExportInventorySearchTests(TestCase):
    def _create_inventory(self, **overrides):
        defaults = {
            "control_number": "CN-EXPORT",
            "office_or_hospital": "Office A",
            "user_name": "User 1",
            "computer_name": "PC-1",
            "assigned_ip": "10.0.0.1",
            "received_by": "Receiver",
            "position": "Staff",
            "date_received": datetime.date(2026, 3, 1),
            "created_at": datetime.date(2026, 3, 19),
            "status": Inventory.Status.ACTIVE,
        }
        defaults.update(overrides)
        return Inventory.objects.create(**defaults)

    def test_export_search_csv_includes_all_components(self):
        inventory = self._create_inventory(control_number="CN-EXPORT")
        other = self._create_inventory(control_number="CN-OTHER", user_name="Other User")

        EquipmentComponent.objects.create(
            inventory=inventory,
            component_name="Processor",
            original_model="Intel i5",
            original_serial="PROC-123",
            replacement_model="Intel i7",
            replacement_serial="PROC-999",
            remarks="Replaced",
        )
        EquipmentComponent.objects.create(
            inventory=inventory,
            component_name="Monitor",
            original_model="Dell 24",
            original_serial="MON-001",
            replacement_model="",
            replacement_serial="",
            remarks="",
        )
        EquipmentComponent.objects.create(
            inventory=other,
            component_name="Keyboard",
            original_model="Logitech",
            original_serial="KBD-777",
            replacement_model="",
            replacement_serial="",
            remarks="",
        )

        url = reverse("inventory:inventory_export_csv") + "?q=CN-EXPORT"
        response = self.client.get(url)

        self.assertEqual(response.status_code, 200)
        text = response.content.decode("utf-8-sig")

        rows = list(csv.reader(io.StringIO(text)))
        self.assertGreaterEqual(len(rows), 3)  # header + 2 components

        header = rows[0]
        self.assertIn("Component", header)
        self.assertIn("Original Model", header)
        self.assertIn("Replacement Serial", header)

        data_rows = rows[1:]
        self.assertEqual(len(data_rows), 2)

        component_names = {r[6] for r in data_rows}  # Component column
        self.assertEqual(component_names, {"Processor", "Monitor"})

        self.assertNotIn("CN-OTHER", text)
        self.assertNotIn("Keyboard", text)

    def test_export_search_csv_includes_inventory_without_components(self):
        self._create_inventory(control_number="CN-NOCOMP")

        url = reverse("inventory:inventory_export_csv") + "?q=CN-NOCOMP"
        response = self.client.get(url)

        self.assertEqual(response.status_code, 200)
        text = response.content.decode("utf-8-sig")

        rows = list(csv.reader(io.StringIO(text)))
        self.assertEqual(len(rows), 2)  # header + 1 inventory row

        self.assertEqual(rows[1][0], "CN-NOCOMP")
        self.assertEqual(rows[1][6], "")  # Component blank

    def test_export_search_excel_includes_all_components_when_available(self):
        inventory = self._create_inventory(control_number="CN-XLSX")
        EquipmentComponent.objects.create(
            inventory=inventory,
            component_name="Processor",
            original_model="Intel i5",
            original_serial="PROC-123",
            replacement_model="",
            replacement_serial="",
            remarks="",
        )
        EquipmentComponent.objects.create(
            inventory=inventory,
            component_name="Monitor",
            original_model="Dell 24",
            original_serial="MON-001",
            replacement_model="",
            replacement_serial="",
            remarks="",
        )

        url = reverse("inventory:inventory_export_excel") + "?q=CN-XLSX"
        response = self.client.get(url)

        try:
            from openpyxl import load_workbook
        except ImportError:
            load_workbook = None

        if load_workbook is None:
            self.assertEqual(response.status_code, 500)
            return

        self.assertEqual(response.status_code, 200)
        workbook = load_workbook(io.BytesIO(response.content))
        sheet = workbook.active

        self.assertEqual(sheet.title, "Inventory Search")

        values = list(sheet.values)
        self.assertEqual(len(values), 3)  # header + 2 components

        header = list(values[0])
        self.assertIn("Component", header)

        component_names = {row[6] for row in values[1:]}
        self.assertEqual(component_names, {"Processor", "Monitor"})


class OfficeScopedInventoryAccessTests(TestCase):
    def setUp(self):
        self.user_model = get_user_model()
        self.office_a_user = self.user_model.objects.create_user(username="office_a", password="password")
        self.office_b_user = self.user_model.objects.create_user(username="office_b", password="password")

        UserProfile.objects.create(user=self.office_a_user, office_or_hospital="Office A")
        UserProfile.objects.create(user=self.office_b_user, office_or_hospital="Office B")

        self.inventory_a = Inventory.objects.create(
            control_number="CN-A",
            office_or_hospital="Office A",
            user_name="User A",
            computer_name="PC-A",
            assigned_ip="10.0.0.10",
            received_by="Receiver",
            position="Staff",
            date_received=datetime.date(2026, 3, 1),
            created_at=datetime.date(2026, 3, 19),
            status=Inventory.Status.ACTIVE,
        )
        self.inventory_b = Inventory.objects.create(
            control_number="CN-B",
            office_or_hospital="Office B",
            user_name="User B",
            computer_name="PC-B",
            assigned_ip="10.0.0.20",
            received_by="Receiver",
            position="Staff",
            date_received=datetime.date(2026, 3, 1),
            created_at=datetime.date(2026, 3, 19),
            status=Inventory.Status.ACTIVE,
        )

    def test_user_sees_only_inventory_from_their_office(self):
        self.client.force_login(self.office_a_user)

        response = self.client.get(reverse("inventory:inventory_list"))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, self.inventory_a.control_number)
        self.assertNotContains(response, self.inventory_b.control_number)

    def test_user_without_delete_permission_does_not_see_delete_action(self):
        self.client.force_login(self.office_a_user)

        response = self.client.get(reverse("inventory:inventory_list"))

        self.assertEqual(response.status_code, 200)
        self.assertNotContains(response, "Delete")

    def test_staff_user_sees_all_inventory(self):
        staff_user = self.user_model.objects.create_user(username="staff_admin", password="password", is_staff=True)
        self.client.force_login(staff_user)

        response = self.client.get(reverse("inventory:inventory_list"))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, self.inventory_a.control_number)
        self.assertContains(response, self.inventory_b.control_number)


class InventoryPrintViewTests(TestCase):
    def _create_inventory(self, **overrides):
        defaults = {
            "control_number": "CN-PRINT",
            "office_or_hospital": "Office A",
            "user_name": "User 1",
            "computer_name": "PC-1",
            "assigned_ip": "10.0.0.1",
            "received_by": "Receiver",
            "position": "Staff",
            "date_received": datetime.date(2026, 3, 1),
            "created_at": datetime.date(2026, 3, 19),
            "status": Inventory.Status.ACTIVE,
        }
        defaults.update(overrides)
        return Inventory.objects.create(**defaults)

    def test_print_view_renders_inventory_details_for_printing(self):
        inventory = self._create_inventory(control_number="CN-PRINT")
        EquipmentComponent.objects.create(
            inventory=inventory,
            component_name="Processor",
            original_model="Intel i5",
            original_serial="PROC-123",
            replacement_model="",
            replacement_serial="",
            remarks="",
        )

        user = get_user_model().objects.create_user(username="print_user", password="password")
        UserProfile.objects.create(user=user, office_or_hospital="Office A")
        self.client.force_login(user)

        response = self.client.get(reverse("inventory:inventory_print", args=[inventory.pk]))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Inventory Print")
        self.assertContains(response, "CN-PRINT")
        self.assertContains(response, "Processor")


class ReportsComponentPaginationTests(TestCase):
    def _create_inventory(self, **overrides):
        defaults = {
            "control_number": "CN-REPORT",
            "office_or_hospital": "Office A",
            "user_name": "User 1",
            "computer_name": "PC-1",
            "assigned_ip": "10.0.0.1",
            "received_by": "Receiver",
            "position": "Staff",
            "date_received": datetime.date(2026, 3, 1),
            "created_at": datetime.date(2026, 3, 19),
            "status": Inventory.Status.ACTIVE,
        }
        defaults.update(overrides)
        return Inventory.objects.create(**defaults)

    def test_reports_does_not_render_equipment_component_records(self):
        inventory = self._create_inventory()

        EquipmentComponent.objects.create(
            inventory=inventory,
            component_name="Processor",
            original_model="Model",
            original_serial="S-000",
            replacement_model="",
            replacement_serial="",
            remarks="",
        )

        response = self.client.get(reverse("inventory:reports"))
        self.assertEqual(response.status_code, 200)
        self.assertNotContains(response, "Equipment Components (Row Format)")
        self.assertNotContains(response, "S-000")
